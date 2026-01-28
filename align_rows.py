import pandas as pd
from openpyxl import load_workbook

file_path = "/Users/mike/Documents/running/2026/swimming-program/RottoTraining MARLENE New New.xlsx"
sheet_name = "Program"

try:
    # Based on the previous output:
    # 'week' seems to be picking up dates (e.g., 2025-10-12) instead of simple numbers (1, 2, 3)
    # The 'phase' and 'target_km' values seem misaligned or shifted.
    # The 'saturday_focus' field looks like it has some of the phase names ("Recovery", "Taper").

    # Re-inspecting the dataframe dump from earlier:
    # Row 39: "Week Commencing", then dates: 2025-10-05, 2025-10-12...
    # Row 40: "Week Number", then numbers: 1, 2, 3...
    # Row 46: "Distance Swum (kms) - Target", values: 9, 10, 10...
    # Row 47: Phase descriptions? "Speed/Sprints (Pool)", "Recovery", "Form/Open Water..."

    # It seems my previous script was slightly off in row/column indices or interpretation.
    # Let's try to map the columns more precisely.
    # The dates start at index 2 (Column C) in the dataframe dump (which is 0-indexed).
    # So 0=A, 1=B, 2=C.

    # Let's grab the raw lists from the rows and print them aligned.

    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # Get rows 39, 40, 46, 47, 48
    # Note: openpyxl rows are 1-based. So row 39 in pandas is row 40 in openpyxl?
    # Let's check the values of the first few cells in these rows to confirm identity.

    def get_row_values(row_idx):
        return [cell.value for cell in ws[row_idx]]

    # Print first few items of candidate rows to confirm which is which
    for r in range(38, 50):
        vals = get_row_values(r)
        # compact print
        print(f"Row {r}: {vals[:5]}")

except Exception as e:
    print(f"Error inspecting rows: {e}")
