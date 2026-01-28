import pandas as pd
from openpyxl import load_workbook
import json
import datetime


def json_serial(obj):
    if isinstance(obj, (datetime.datetime, datetime.date)):
        return obj.isoformat()
    raise TypeError("Type %s not serializable" % type(obj))


file_path = "/Users/mike/Documents/running/2026/swimming-program/RottoTraining MARLENE New New.xlsx"
sheet_name = "Program"

try:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # --- 1. Extract The Base Schedule (Weeks, Dates, Targets, Phases) ---
    schedule = []

    # We established these indices:
    row_date = 40
    row_week_num = 41
    row_target_km = 47
    row_phase = 48
    row_sat_focus = (
        49  # Just in case, though it looked empty, row 48 had merged values.
    )

    # Iterate columns B (2) to V (22)
    for col_idx in range(2, 23):
        # Handle Merged Cells for Phase
        phase_cell = ws.cell(row=48, column=col_idx)
        phase_val = phase_cell.value

        # Look for merged value if None
        if phase_val is None:
            for merged_range in ws.merged_cells.ranges:
                if phase_cell.coordinate in merged_range:
                    top_left = ws.cell(
                        row=merged_range.min_row, column=merged_range.min_col
                    )
                    phase_val = top_left.value
                    break

        # Saturday specific notes?
        # In the sample output:
        # Col 4 (Oct 26): Row 48 has "Recovery".
        # Col 5 (Nov 02): Row 48 has "Form/Open Water...".
        # Wait, the previous extraction showed these in "phase_focus".
        # Let's see if there is a separate "Saturday Focus" row.
        # Row 39 had "Saturday" in Col E (5).
        # Row 48 seems to contain the high-level focus for the week/phase.

        # Let's try to capture the specific instructions found in rows 54-62.
        # These seemed to be column-specific notes.
        notes = []
        for r in range(54, 63):
            val = ws.cell(row=r, column=col_idx).value
            if val and isinstance(val, str):
                notes.append(val)

        # Structure the data
        week_data = {
            "week": ws.cell(row=41, column=col_idx).value,
            "start_date": ws.cell(row=40, column=col_idx).value,
            "target_km": ws.cell(row=47, column=col_idx).value,
            "phase": phase_val,
            "notes": notes,
        }
        schedule.append(week_data)

    # --- 2. Construct the Final JSON Object ---

    final_output = {
        "program_name": "Rottnest Channel Swim 2025-26",
        "description": "Training Schedule Charts",
        "schedule": schedule,
    }

    print(json.dumps(final_output, default=json_serial, indent=2))

except Exception as e:
    print(f"Error generating final JSON: {e}")
