import pandas as pd
from openpyxl import load_workbook

file_path = "/Users/mike/Documents/running/2026/swimming-program/RottoTraining MARLENE New New.xlsx"
sheet_name = "Program"

try:
    # Use openpyxl to inspect cell values directly for a section of the sheet
    # Based on the previous output, the dates and week numbers are in rows 39 and 40
    # Let's inspect rows 50 to 100 to see where the actual workout details are
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    print("Inspecting rows 50-100:")
    for row in ws.iter_rows(min_row=50, max_row=100, max_col=5):
        row_data = [cell.value for cell in row]
        # Only print rows that aren't entirely None
        if any(row_data):
            print(f"Row {row[0].row}: {row_data}")

except Exception as e:
    print(f"Error inspecting sheet '{sheet_name}': {e}")
