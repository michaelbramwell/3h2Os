import pandas as pd
from openpyxl import load_workbook

file_path = "/Users/mike/Documents/running/2026/swimming-program/RottoTraining MARLENE New New.xlsx"
sheet_name = "Program"

try:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # Based on the initial "head" dump, row 46 has "Distance Swum (kms) - Target"
    # and row 47 has phases like "Speed/Sprints (Pool)".
    # The columns correspond to weeks. Column 1 is "Week Commencing", 2 is the first date, etc.
    # It seems the actual workout descriptions might be missing or located elsewhere,
    # OR this sheet is just a high-level volume summary.

    # Let's search the whole sheet for keywords like "Tuesday", "Thursday", "Saturday"
    # or typical swim terms like "warm up", "main set", "drill".

    print("Searching for keywords in 'Program' sheet:")
    found_keywords = False
    for row in ws.iter_rows(max_row=200):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_lower = cell.value.lower()
                if any(
                    x in val_lower
                    for x in [
                        "tuesday",
                        "thursday",
                        "saturday",
                        "warm up",
                        "warmup",
                        "main set",
                        "drill",
                        "cool down",
                    ]
                ):
                    print(f"Found '{cell.value}' at {cell.coordinate}")
                    found_keywords = True

    if not found_keywords:
        print("No specific workout keywords found in the first 200 rows.")

except Exception as e:
    print(f"Error inspecting sheet '{sheet_name}': {e}")
