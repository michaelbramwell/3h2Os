import pandas as pd
from openpyxl import load_workbook

file_path = "/Users/mike/Documents/running/2026/swimming-program/RottoTraining MARLENE New New.xlsx"
sheet_name = "Program"

try:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    print("--- SCANNING FOR WORKOUT DETAILS (Rows 50-100) ---")

    # We suspect workout details are aligned with the phase columns.
    # Let's inspect the text in rows 50-100 for columns B (2) through V (22)
    # to see if we can find the detailed descriptions.

    # We will group non-empty cells by column to see if they form vertical blocks of instructions

    column_content = {}

    for col_idx in range(2, 23):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        cells_with_text = []
        for row_idx in range(50, 100):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and isinstance(val, str) and len(val.strip()) > 0:
                cells_with_text.append(f"R{row_idx}: {val.strip()}")

        if cells_with_text:
            column_content[col_idx] = cells_with_text

    for col, texts in column_content.items():
        print(f"\n--- Column {col} Content ---")
        for t in texts:
            print(t)

except Exception as e:
    print(f"Error inspecting workout details: {e}")
