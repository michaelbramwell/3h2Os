import pandas as pd
from openpyxl import load_workbook
import json
import datetime


def json_serial(obj):
    """JSON serializer for objects not serializable by default json code"""
    if isinstance(obj, (datetime.datetime, datetime.date)):
        return obj.isoformat()
    raise TypeError("Type %s not serializable" % type(obj))


file_path = "/Users/mike/Documents/running/2026/swimming-program/RottoTraining MARLENE New New.xlsx"
sheet_name = "Program"

try:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # Correct Row Mappings based on previous output:
    # Row 40: Dates (Week Commencing)
    # Row 41: Week Number
    # Row 47: Distance Swum (kms) - Target
    # Row 48: Phase / Focus (Merged cells might be tricky here)
    # Row 49: Saturday Specifics? (Wait, previous dump showed Row 49 was empty, but Row 48 had 'Recovery')
    # Let's look at Row 49 again later if needed. Row 50 had 'COACH BK'.

    # We need to extract the structure.
    # Columns B (idx 2) to V (idx 22) seem to cover the weeks.

    weeks_data = []

    # Iterate columns 2 to 22 (inclusive)
    for col_idx in range(2, 23):
        # Week Commencing Date
        date_cell = ws.cell(row=40, column=col_idx)
        start_date = date_cell.value

        # Week Number
        week_num = ws.cell(row=41, column=col_idx).value

        # Target Distance
        target_km = ws.cell(row=47, column=col_idx).value

        # Phase / Focus
        # Note: Merged cells in Excel often only have the value in the top-left cell.
        # We need to handle this. If cell value is None, it might be part of a merge.
        phase_cell = ws.cell(row=48, column=col_idx)
        phase = phase_cell.value

        # Check if this cell is part of a merged range
        is_merged = False
        for merged_range in ws.merged_cells.ranges:
            if phase_cell.coordinate in merged_range:
                # If we are not the top-left, we inherit the value
                # But openpyxl's data_only=True might not fill it?
                # Actually, we need to find the top-left value of the range
                top_left_cell = ws.cell(
                    row=merged_range.min_row, column=merged_range.min_col
                )
                phase = top_left_cell.value
                is_merged = True
                break

        # Saturday specific notes?
        # Let's check row 49 (which was empty in sample) and row 50 (Coach BK)
        # Maybe row 39 (Saturday)? It had "Saturday" in column E.

        weeks_data.append(
            {
                "week_number": week_num,
                "start_date": start_date,
                "target_distance_km": target_km,
                "phase_focus": phase,
            }
        )

    # Print structured result
    print("--- EXTRACTED SWIMMING PROGRAM STRUCTURE ---")
    print(json.dumps(weeks_data, default=json_serial, indent=2))

except Exception as e:
    print(f"Error extracting program: {e}")
