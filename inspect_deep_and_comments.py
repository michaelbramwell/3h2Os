import pandas as pd
from openpyxl import load_workbook

file_path = "/Users/mike/Documents/running/2026/swimming-program/RottoTraining MARLENE New New.xlsx"

try:
    wb = load_workbook(file_path, data_only=True)

    # 1. Check "Port to Pub" sheet
    if "Port to Pub" in wb.sheetnames:
        print("\n--- Inspecting 'Port to Pub' Sheet (First 20 rows) ---")
        ws_p2p = wb["Port to Pub"]
        for r in range(1, 21):
            vals = [c.value for c in ws_p2p[r]]
            if any(vals):
                print(f"Row {r}: {vals[:10]}")

    # 2. Check "Program" sheet deeper
    ws_prog = wb["Program"]
    print("\n--- Scanning 'Program' sheet deeper (Rows 100-300) ---")
    found_deep_content = False
    for r in range(100, 301):
        vals = [c.value for c in ws_prog[r]]
        # Check if row has significant content (more than just empty/none)
        # Convert to string to check for keywords
        row_str = " ".join([str(v) for v in vals if v is not None]).lower()
        if any(
            k in row_str
            for k in ["warm up", "warmup", "main set", "drill", "x 100", "x 50"]
        ):
            print(f"Row {r}: {vals[:5]}...")
            found_deep_content = True
            if not found_deep_content:  # Just print a few to confirm
                break

    if not found_deep_content:
        print("No workout keywords found in deep scan.")

    # 3. Check for Comments in "Program" sheet
    print("\n--- Checking for Cell Comments in 'Program' Sheet ---")
    comments_found = 0
    for row in ws_prog.iter_rows():
        for cell in row:
            if cell.comment:
                print(f"Comment at {cell.coordinate}: {cell.comment.text}")
                comments_found += 1
                if comments_found > 5:
                    print("... (stopping after 5 comments)")
                    break
        if comments_found > 5:
            break

    if comments_found == 0:
        print("No comments found.")

except Exception as e:
    print(f"Error inspecting deep/comments: {e}")
