import pandas as pd
from openpyxl import load_workbook

file_path = "/Users/mike/Documents/running/2026/swimming-program/RottoTraining MARLENE New New.xlsx"
sheet_name = "Program"

try:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # We found dates in row 39 (starting index 1). The actual dates start from column 2 (index 1 is Week Commencing text).
    # Let's verify the exact structure of the date row.

    dates = []
    # Row 39 is index 39 in 1-based openpyxl, but let's be safe and iterate
    # The printed dataframe showed row 39 had the dates.

    # Let's extract the weekly volume targets and the phase descriptions.
    # From previous output:
    # Row 39: Dates
    # Row 40: Week Number
    # Row 46: Distance Swum (kms) - Target
    # Row 47: Phase (Speed/Sprints, Recovery, etc.)
    # Row 48: Specific Saturday swim? (CJJ Both Ways, CJJ Fence Swims)

    print("Extracting schedule structure...")

    schedule = []

    # Iterate through columns starting from column C (index 3) which seems to be the first date
    # In openpyxl, columns are 1-based. A=1, B=2, C=3.
    # The previous dataframe output showed dates starting at col index 2 (which is C).

    start_col = 3
    # Find the max column with data
    max_col = ws.max_column

    for col in range(start_col, max_col + 1):
        # Row 40 is typically week number
        week_num = ws.cell(row=40, column=col).value
        date_val = ws.cell(row=39, column=col).value
        target_dist = ws.cell(row=46, column=col).value
        phase = ws.cell(row=47, column=col).value
        saturday_focus = ws.cell(row=48, column=col).value

        if week_num is not None or date_val is not None:
            schedule.append(
                {
                    "week": week_num,
                    "date": str(date_val),
                    "target_km": target_dist,
                    "phase": phase,
                    "saturday_focus": saturday_focus,
                }
            )

    print(f"Found {len(schedule)} weeks.")
    for item in schedule:
        print(item)

except Exception as e:
    print(f"Error extracting schedule: {e}")
